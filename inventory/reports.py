from datetime import date, timedelta
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import F
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import Medicine

class MedicineReportGenerator:
    def __init__(self, user):
        self.user = user
        self.medicines = Medicine.objects.filter(user=user)
    
    def get_inventory_summary(self):
        """Get summary statistics for inventory"""
        total_medicines = self.medicines.count()
        expired_medicines = self.medicines.filter(expiry_date__lt=timezone.now().date()).count()
        expiring_soon = self.medicines.filter(
            expiry_date__lte=timezone.now().date() + timedelta(days=30),
            expiry_date__gt=timezone.now().date()
        ).count()
        low_stock = self.medicines.filter(quantity__lte=F('low_stock_threshold')).count()
        
        return {
            'total_medicines': total_medicines,
            'expired_medicines': expired_medicines,
            'expiring_soon': expiring_soon,
            'low_stock': low_stock,
        }
    
    def generate_excel_report(self, report_type='all'):
        """Generate Excel report with formatting"""
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Medicine Report - {report_type.title()}"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write header
        headers = [
            'Medicine Name', 'Batch Number', 'Manufacturer', 'Manufacturing Date',
            'Expiry Date', 'Quantity', 'Price per Unit', 'Low Stock Threshold',
            'Description', 'Status', 'Days Until Expiry'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Filter medicines based on report type
        medicines = self._filter_medicines_by_type(report_type)
        
        # Write data
        for row, medicine in enumerate(medicines, 2):
            status = self._get_medicine_status(medicine)
            days_until_expiry = medicine.days_until_expiry
            
            data = [
                medicine.name,
                medicine.batch_number,
                medicine.manufacturer,
                medicine.manufacturing_date,
                medicine.expiry_date,
                medicine.quantity,
                medicine.price_per_unit,
                medicine.low_stock_threshold,
                medicine.description or '',
                status,
                days_until_expiry
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                
                # Color coding based on status
                if status == 'Expired':
                    cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                elif status == 'Expiring Soon':
                    cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
                elif status == 'Low Stock':
                    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="medicine_report_{report_type}_{date.today()}.xlsx"'
        
        wb.save(response)
        return response
    
    def generate_pdf_report(self, report_type='all'):
        """Generate PDF report"""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="medicine_report_{report_type}_{date.today()}.pdf"'
        
        # Create PDF document
        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Add title
        title = Paragraph(f"Medicine Inventory Report - {report_type.title()}", title_style)
        elements.append(title)
        
        # Add summary
        summary = self.get_inventory_summary()
        summary_data = [
            ['Total Medicines', summary['total_medicines']],
            ['Expired Medicines', summary['expired_medicines']],
            ['Expiring Soon (30 days)', summary['expiring_soon']],
            ['Low Stock Items', summary['low_stock']],
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 1*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # Filter medicines based on report type
        medicines = self._filter_medicines_by_type(report_type)
        
        if medicines:
            # Prepare data for table
            data = [['Name', 'Batch', 'Manufacturer', 'Expiry', 'Quantity', 'Status']]
            
            for medicine in medicines:
                status = self._get_medicine_status(medicine)
                data.append([
                    medicine.name,
                    medicine.batch_number,
                    medicine.manufacturer,
                    medicine.expiry_date.strftime('%Y-%m-%d'),
                    str(medicine.quantity),
                    status
                ])
            
            # Create table
            table = Table(data, colWidths=[1.5*inch, 1*inch, 1.5*inch, 1*inch, 0.8*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            elements.append(table)
        else:
            no_data = Paragraph("No medicines found for this report type.", styles['Normal'])
            elements.append(no_data)
        
        # Build PDF
        doc.build(elements)
        return response
    
    def _filter_medicines_by_type(self, report_type):
        """Filter medicines based on report type"""
        if report_type == 'all':
            return self.medicines
        elif report_type == 'expired':
            return self.medicines.filter(expiry_date__lt=timezone.now().date())
        elif report_type == 'expiring_soon':
            return self.medicines.filter(
                expiry_date__lte=timezone.now().date() + timedelta(days=30),
                expiry_date__gt=timezone.now().date()
            )
        elif report_type == 'low_stock':
            return self.medicines.filter(quantity__lte=F('low_stock_threshold'))
        elif report_type == 'active':
            return self.medicines.filter(
                expiry_date__gt=timezone.now().date(),
                quantity__gt=F('low_stock_threshold')
            )
        else:
            return self.medicines
    
    def _get_medicine_status(self, medicine):
        """Get status of medicine"""
        if medicine.is_expired:
            return 'Expired'
        elif medicine.is_expiring_soon:
            return 'Expiring Soon'
        elif medicine.is_low_stock:
            return 'Low Stock'
        else:
            return 'Good' 